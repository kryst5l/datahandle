# -*- coding: utf-8 -*-   
       
import os  
from openpyxl import *
from xpinyin import Pinyin


def get_file_names(file_dir):
	l = []
	for root, dirs, files in os.walk(file_dir):
		for file in files:
			l.append(file)
	return l

def get_pinyin_name(name_list):
	pinyin_name = []
	p = Pinyin()
	for name in sheet_names:
		name = p.get_pinyin(name)
		pinyin_name.append(name)
	return pinyin_name

def read_excel(sheet,col,count):
	if count[0] == '1':
		first = int(count[0]) + 1
	else:
		first = int(count[0]) + 1
	l = []
	while True:
		content = sheet.cell(row = first, column = col + 1).value
		l.append(content)
		if first == int(count[1]) + 2 :
			break
		first += 1
	return l
	

#path = 'C:\\Users\\mobvoi\\Desktop\\0114'
path = '/home/mobvoi/data_collect/CNS/0116/0116'
# wb = load_workbook('陈治超1228.xlsx')
workbook = Workbook()
new_sheet = workbook.active
count = []
f = open('/home/mobvoi/data_collect/CNS/CNS.txt','r')
all_content = f.readlines()

for file in get_file_names(path):
	file = file.split('.')[0]
	index_list = file.split('_')[-2].split('to')
	conten = all_content[int(index_list[0])-1:int(index_list[1])]
	new_sheet.append([file,','.join(conten).replace('\n','')])
workbook.save('type.xlsx')
f.close()